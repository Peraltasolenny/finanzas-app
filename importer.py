"""Lectura de estados de cuenta para importar transacciones.

Soporta:
  - PDF digital (con texto), usando pdfplumber.
  - CSV (separado por comas o punto y coma).
  - Excel (.xlsx), usando openpyxl.

Devuelve siempre una lista de dicts:
    {"tx_date": date, "description": str, "amount": float, "type": "income"|"expense"}
El usuario revisa y confirma antes de que se guarde nada.
"""
import csv
import io
import re
from datetime import datetime, date

# Extensiones aceptadas en el formulario de carga.
ALLOWED_EXTENSIONS = {".pdf", ".csv", ".xlsx", ".xls"}

# Palabras que sugieren que un movimiento es un ingreso/pago (no un gasto).
_INCOME_HINTS = ("pago a tarjeta", "pago recibido", "payment", "abono",
                 "deposito", "depósito", "transferencia recibida", "credito",
                 "crédito", "reembolso", "devolucion", "devolución", "salario",
                 "nomina", "nómina", "cashback")

_MONEY_TOKEN = r"(?:RD\$|US\$|DOP|USD|\$)?\s*-?[\d.,]+\d"
# Línea típica de estado de cuenta: una o dos fechas, descripción y un monto al final.
_PDF_LINE_RE = re.compile(
    r"^\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"          # fecha 1
    r"(?:\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4})?"         # fecha 2 (opcional, se ignora)
    r"\s+(.+?)\s+"                                    # descripción (no glotona)
    r"(?:RD\$|US\$|DOP|USD|\$)\s*([\d.,]+\.\d{2})\s*$"  # monto final
)

_DATE_FORMATS = ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
                 "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y")


def _parse_date(text):
    text = (text or "").strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(text):
    """Convierte '20,369.40', 'RD$ 1.862,78' o '-1234.5' a float (valor absoluto)."""
    if text is None:
        return None
    s = re.sub(r"(RD\$|US\$|DOP|USD|\$)", "", str(text)).strip()
    s = s.replace(" ", "")
    if not s:
        return None
    neg = s.startswith("-") or s.startswith("(")
    s = s.strip("()-")
    # Detecta separador decimal: si hay coma y punto, el último es el decimal.
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):      # formato 1.862,78
            s = s.replace(".", "").replace(",", ".")
        else:                                 # formato 1,862.78
            s = s.replace(",", "")
    elif "," in s:
        # Solo coma: si está a 3 del final es separador de miles, si no, decimal.
        s = s.replace(",", "") if re.search(r",\d{3}$", s) else s.replace(",", ".")
    try:
        val = abs(float(s))
        return val if not neg else val  # guardamos siempre positivo; el tipo define el signo
    except ValueError:
        return None


def _guess_type(description):
    d = (description or "").lower()
    return "income" if any(h in d for h in _INCOME_HINTS) else "expense"


def _clean_desc(text):
    text = re.sub(r"\s+", " ", text or "").strip()
    return text.strip(" -·").strip()


# ----------------- PDF -----------------
def parse_pdf(fileobj):
    import logging
    import pdfplumber  # import perezoso: solo si se usa

    # Silencia advertencias ruidosas de fuentes (no afectan la extracción).
    logging.getLogger("pdfminer").setLevel(logging.ERROR)

    rows = []
    with pdfplumber.open(fileobj) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                m = _PDF_LINE_RE.match(line)
                if not m:
                    continue
                d = _parse_date(m.group(1))
                amount = _parse_amount(m.group(3))
                desc = _clean_desc(m.group(2))
                if d and amount and amount > 0:
                    rows.append({
                        "tx_date": d, "description": desc[:255],
                        "amount": amount, "type": _guess_type(desc),
                    })
    return rows


# ----------------- CSV / Excel -----------------
_COL_DATE = ("fecha", "date", "día", "dia")
_COL_DESC = ("descrip", "concepto", "detalle", "description", "memo", "referencia", "comercio")
_COL_AMOUNT = ("monto", "importe", "amount", "valor", "debito", "débito", "credito", "crédito", "cargo")


def _match_col(header, keywords):
    for i, h in enumerate(header):
        hl = str(h or "").strip().lower()
        if any(k in hl for k in keywords):
            return i
    return None


def _rows_from_table(table):
    """table: lista de filas (listas). Detecta encabezado y extrae movimientos."""
    table = [r for r in table if any(str(c).strip() for c in r)]
    if not table:
        return []

    # Busca la fila de encabezado (la que mencione fecha + monto).
    header_idx = 0
    for i, row in enumerate(table[:10]):
        if _match_col(row, _COL_DATE) is not None and _match_col(row, _COL_AMOUNT) is not None:
            header_idx = i
            break
    header = table[header_idx]
    ci_date = _match_col(header, _COL_DATE)
    ci_desc = _match_col(header, _COL_DESC)
    ci_amount = _match_col(header, _COL_AMOUNT)

    rows = []
    for row in table[header_idx + 1:]:
        if ci_date is None or ci_amount is None or ci_date >= len(row) or ci_amount >= len(row):
            continue
        d = _parse_date(str(row[ci_date]))
        amount = _parse_amount(row[ci_amount])
        desc = _clean_desc(str(row[ci_desc])) if ci_desc is not None and ci_desc < len(row) else ""
        if d and amount and amount > 0:
            rows.append({
                "tx_date": d, "description": desc[:255],
                "amount": amount, "type": _guess_type(desc),
            })
    return rows


def parse_csv(fileobj):
    raw = fileobj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig", errors="replace")
    # Detecta delimitador (coma o punto y coma).
    sample = raw[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
    return _rows_from_table(list(reader))


def parse_excel(fileobj):
    from openpyxl import load_workbook  # import perezoso

    wb = load_workbook(fileobj, read_only=True, data_only=True)
    ws = wb.active
    table = []
    for row in ws.iter_rows(values_only=True):
        table.append(["" if c is None else c for c in row])
    wb.close()
    return _rows_from_table(table)


# ----------------- dispatcher -----------------
def parse_statement(fileobj, filename):
    """Devuelve (rows, error). rows es lista de dicts; error es str o None."""
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        return [], f"Tipo de archivo no soportado: {ext or 'desconocido'}. Usa PDF, CSV o Excel."
    try:
        if ext == ".pdf":
            rows = parse_pdf(fileobj)
        elif ext == ".csv":
            rows = parse_csv(fileobj)
        else:  # .xlsx / .xls
            rows = parse_excel(fileobj)
    except Exception as exc:  # noqa: BLE001 — queremos mostrar el error al usuario
        return [], f"No se pudo leer el archivo: {exc}"

    if not rows:
        return [], ("No se detectaron movimientos. Si es un PDF escaneado (imagen), "
                    "el texto no se puede extraer; exporta el movimiento en CSV/Excel desde tu banco.")
    # Ordena por fecha ascendente.
    rows.sort(key=lambda r: r["tx_date"])
    return rows, None
